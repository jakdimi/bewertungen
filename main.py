from openpyxl import Workbook, load_workbook

group_name = input("Give the name of your group: \n")
target_path = input("Where should the result be saved? (If left blank this will default to ./result/): \n")
if not target_path:
    target_path = "./result/"

source_grading_file = input("Give the file of the gradings (If left blank this defaults to ./resources/grading.xlsx): \n")
if not source_grading_file:
    source_grading_file = "./resources/grading.xlsx"
source_participants_file = input("give the file of the participants (If left blank this will default to ./resources/participants.xlsx): \n")
if not source_participants_file:
    source_participants_file = "./resources/participants.xlsx"

output = Workbook()
out_sheet = output.active
grading = load_workbook(filename=source_grading_file)
grading_sheet = grading.active
participants = load_workbook(filename=source_participants_file)
participants_sheet = participants.active

last_empty_row = 2


def next_column(column):
    nr_of_Z = 0
    for c in column[::-1]:
        if c == "Z":
            nr_of_Z += 1
        else:
            break
    
    if nr_of_Z == len(column):
        return ''.join(['A' for i in range(len(column) + 1)])
    
    out_list = list(column)
    out_list[-nr_of_Z - 1] = chr(ord(out_list[-nr_of_Z - 1]) + 1)
    for i in range(nr_of_Z):
        out_list[-i-1] = 'A'

    return ''.join(out_list)


def row_of_email(email):
    row = 2
    while grading_sheet[f"F{row}"].value:
        current_email = grading_sheet[f"F{row}"].value
        if current_email == email:
            return row
        row += 1
    return -1


def get_row(sheet, row_nr):
    row = []
    column = "A"
    while sheet[f"{column}{row_nr}"].value is not None:
        row.append(sheet[f"{column}{row_nr}"].value)
        column = next_column(column)
    return row


def get_column_signature(n):
    out = "A"
    for i in range(n):
        out = next_column(out)
    return out


def set_row(sheet, row, row_nr):
    column = "A"
    for entry in row:
        sheet[f"{column}{row_nr}"] = entry
        column = next_column(column)


def print_disputes():
    grades_keys = get_row(grading_sheet, 1)

    for i in range(1, 20):
        group = []
        disputes = []
        row = 2
        while participants_sheet[f"E{row}"].value:
            email = participants_sheet[f"E{row}"].value
            groups = participants_sheet[f"F{row}"].value
            if groups is not None:
                if group_name in groups and f"AG{i}" in groups:
                    grading_row = row_of_email(email)
                    if grading_row != -1:
                        grades_row = get_row(grading_sheet, row_of_email(email))
                        group.append(grades_row)
            
            row += 1
        
        if len(group) != 0:
            for j in range(len(group[0])):
                if "Aufgabe" in grades_keys[j] and "Abgabe" in grades_keys[j] and "Blatt 1" not in grades_keys[j]:
                    value = group[0][j]
                    for participant in group:
                        if participant[j] != value:
                            disputes.append(j)
                            break
            
            for dispute in disputes:
                print(f"Dispute in AG{i}: {grades_keys[dispute]}")
                for participant in group:
                    print(f"{participant[0]} {participant[1]}: {participant[dispute]}")
                print()


def make_output():
    global last_empty_row
    grades_keys = get_row(grading_sheet, 1)

    out_keys = ["Vorname", 
                "Nachname", 
                "Pkt. Gesamt", 
                ">= 6", "< 3", 
                "Videos", 
                "bestandene Aufgaben"]

    set_row(out_sheet, out_keys, 1)

    row = 2
    while participants_sheet[f"E{row}"].value:
        email = participants_sheet[f"E{row}"].value
        groups = participants_sheet[f"F{row}"].value
        
        is_participant_in_group = False
        if groups is not None:
            if group_name in groups:
                grading_row = row_of_email(email)
                if grading_row != -1:
                    is_participant_in_group = True
        
        if is_participant_in_group:
            row_content = []
            grades_row = get_row(grading_sheet, row_of_email(email))
            row_content.append(grades_row[0])
            row_content.append(grades_row[1])
            row_content.append(grades_row[6])

            excercise = 0
            exercise_failed = 0
            video = 0
            for i, value in enumerate(grades_row[7:], start=7):
                column_is_loesung = False
                column_is_abgabe= False
                if "Aufgabe" in grades_keys[i] and value != "-":
                    if "Lösung" in grades_keys[i]:
                        column_is_loesung = True
                    if "Abgabe" in grades_keys[i]:
                        column_is_abgabe = True
                
                if column_is_loesung and int(value) >= 1:
                    video += 1
                
                if column_is_abgabe and int(value) >= 6:
                    excercise += 1

                if column_is_abgabe and int(value) < 3:
                    exercise_failed += 1

                if column_is_abgabe and int(value) == 2:
                    print(
                        f"{grades_row[0]} {grades_row[1]}, \
                            {grades_keys[i]}: {grades_row[i]}, \
                            {grades_keys[i+3]}: {grades_row[i+3]}"
                        )
            
            kurs_gesamt = grades_row[7]

            row_content.append(excercise)
            row_content.append(exercise_failed)
            row_content.append(video)
            row_content.append(kurs_gesamt)

            set_row(out_sheet, row_content, last_empty_row)
            last_empty_row += 1
        row += 1
    
    output.save(filename=f"{target_path}out.xlsx")


def main():
    make_output()


if __name__ == "__main__":
    main()

